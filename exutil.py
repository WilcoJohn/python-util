# -*- coding: utf-8 -*-
"""
Module for Excel cell processing, string similarity, and pattern-based file filtering.

Created on Tue Aug  5 07:08:28 2025

@author: WilcoSievers
"""

from openpyxl.worksheet.worksheet import Worksheet # for type hinting
from numbers import Number# for type hinting


def exc_coord_to_rc(cell_Coord : str) -> (int, int): 
    """
    Converts an Excel cell coordinate (e.g., "A1", "AA10") to (row, column) integers.

     Args:
        cell_Coord: Excel-style cell coordinate string.

    Returns:
        A tuple (row, column) with integer indices
     
    Ex.
    A3      ->   row = 3    ,  column = 1
    AA33    ->   row = 33   ,  column = 77
    B4      ->   row = 4    ,  column = 2
    BAA501  ->   row = 501  ,  column = 53   
    """
    from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
    # Error/ type handling.
    if not cell_Coord:
        raise ValueError(f"Invalid cell coordinate give: {cell_Coord}")
    if not isinstance(cell_Coord, str):
        raise TypeError(f"excelCordToRC(cell_Coord) requires string type to convert, gave : {type(cell_Coord)} type")
    
    col_letter, row = coordinate_from_string(cell_Coord);    # get row value
    col = column_index_from_string(col_letter);              # get column value
    return row, col;

def filter_files(path_list:  str | list[str] = r"./", 
                pattern: str = "*", 
                matchPattern: bool = True) -> list[str]:  
    """
    Filters files by a Unix-style wildcard pattern.
    
    Args:
        path_list = A directory path (str) to search recursively, or a list/tuple of file paths.
        pattern   = Patterns are Unix shell style:
    
                    *       matches everything
                    ?       matches any single character
                    [seq]   matches any character in seq
                    [!seq]  matches any char not in seq
    
    matchPattern = If 'True', include files that match the pattern; if False, exclude them.

    Returns:
        A sorted list of filter str

        
    """ 
    import os
    from fnmatch import fnmatch

    results = [];
    
    if isinstance(path_list, str):                                       # if path is str, search for files to path
        for root, _, files in os.walk(path_list):                        # walk through 
            for fi in files:
                full_path = os.path.join(root, fi);
                if (fnmatch(fi, pattern) == matchPattern):       # if file match pattern then append
                    results.append(full_path)
                    
    elif isinstance(path_list,(list, tuple)):                            # filter iteratible list for relevant patter
        for item in path_list:
             if (fnmatch(item, pattern) == matchPattern):          # if str matches pattern then append
                    results.append(full_path)
                 
    else:
        raise TypeError(f"Unrecognised data type for path_list feed var: {type(path_list)}, expected str or tuple of strings.");
         
    return sorted(results);


def is_similar(target : str = '', 
               test : str | list[str] = '', 
               threshold : float = 0.8, 
               return_similarrity_score : bool = False) -> bool | tuple[str, float]:
    """
    Determine if the given 'test' text is similar to the 'target' text

    Args:
        target = Target str to compare against
        test   = Iteratable of str or list to compare to target
        threshold = Similarity criteria that needs to be adhered to
        return_similarrity_score = Returns the most similar str that overcomes the threshold 

    Returns:
        Bool if value similarity is above 'threshold'
        Most similar value if string is above threshold and 'return_similarrity_score' is 'True'
    """
    import numpy
    from difflib import  SequenceMatcher
    
    # Target value error/ type handling
    if not isinstance(target, str):        # check if target is in string format, 
        try:
            target = str(target)
        except Exception:
            raise TypeError(f"Target must be convertible to string, got {type(target)}")
    target = target.strip();       

    
    # Test variable error/ type handling
    # Check format type of test value. If test is of list-type, 
    # or numeric type, convert to str and strip is to lower case.                                               
    if isinstance(test, (list, numpy.ndarray)):
        test = [str(ti).strip()for ti in test];   
    elif isinstance(test, str):
        test = test.strip();       # convert to str
    elif isinstance(test, Number):
        test = str(test).strip();
    else:
        raise TypeError(f"Test value must be str, number, or iterable of strings. Got {type(test)}")

    
    # Case 1: simple string-to-string comparison
    if isinstance(test, str):                   # return true if string is similar
        similarity = SequenceMatcher(None, target, test).ratio();
        if return_similarrity_score:
            return (test, similarity) if similarity >= threshold else (None, None);
        return similarity >= threshold;

    # Case 2: iterable of strings and return most similar
    if isinstance(test, (list, numpy.ndarray, tuple)) and return_similarrity_score:
        valid_str_list = [];
        valid_str_score = [];
        
        for s_i in test:
            similarity = SequenceMatcher(None, target, s_i).ratio()
            if (similarity >= threshold):
                valid_str_list.append(s_i);
                valid_str_score.append(similarity);
                
        if valid_str_list:
            return numpy.array(valid_str_list), numpy.array(valid_str_score);
        return (None, None)

    # Case 3: iterable of strings without return_similarrity_score → return True if any matchopen
    if isinstance(test, (list, tuple, numpy.ndarray)):
        return any(SequenceMatcher(None, target, s).ratio() >= threshold for s in test)
    
    # Should never reach here
    raise RuntimeError(
        f"Unexpected logic path.\nTarget: {target}\nTest: {test}\n"
        f"Threshold: {threshold}\nReturn most similar: {return_similarrity_score}")






def count_significant_digits(number : Number = 1.0):
    """
    Counts the number of significant digits in a number represented as a string.
    Handles trailing zeros after a decimal point as significant.
    """
    
    from decimal import Decimal  

    if isinstance(number, Number):       # first try and convert number to str, Python can do this fairly well with no problem.
        number_str = str(number);
    elif not isinstance(number, str):           # if not number, convert water the data type, catch the error
        number_str = str(number);
    else:
        number_str = number;

    float(number_str); # try cinverting values, if number contains number then a value error will be raised
    
    d = Decimal(number_str)
    # Remove leading zeros and handle scientific notation
    normalized_str = str(d.normalize()) 

    # Handle cases like "1.000" where normadlize() might remove trailing zeros
    # if they are considered insignificant in the standard decimal representation.
    # We want to preserve them if they were explicitly given.
    if '.' in number_str and not normalized_str.endswith('0') and number_str.endswith('0'):
        # Count explicit trailing zeros after the decimal point
        trailing_zeros = len(number_str) - len(number_str.rstrip('0'))
        return len(normalized_str.replace('.', '')) + trailing_zeros
    
    return len(normalized_str.replace('.', '').replace('-', '')) # Remove decimal point and sign for counting

def search_excl_val(sheet : Worksheet, 
                    search_targets, 
                    threshold : float = 0.8,
                    return_first_hit : bool = True,
                    return_all_vals : bool = False):
    """
    Search through an excel worksheet to find coordinates equal to search val

    Args:
        sheet = openpyxl sheet object to search through
        search_targets = iterable to search through. If the first value ( in position '0') is not a hit, continue through next iterable. Iterables need to be of same data type.
        threshold = Similarity threshold to search through
        return_first_hit = Return coordinate of first wxcwl cell equal to select value
    Return:
        A list of cell coordinates in excel 'A3' standard.
    """

    import numbers
    import datetime
    from numpy import round
    
    # Error handling: check if sheet is of correct data-type
    if not isinstance(sheet, Worksheet):   
        raise TypeError(f"Incorrect variable type given for sheet '{sheet}' - {type(sheet)}, expected {type(Worksheet)}")

     # Error handling: check if check if search targets are all the same data typee
    search_type = type(search_targets[0])                              
    if not all(isinstance(target_i, search_type) for target_i in search_targets): 
        raise TypeError(f"Search targets not same data type : {search_targets}")
    
    similar_target_hits = {'Coordinate' : [], 'Value' : [] };
    equal_targets = {'Coordinate' : [], 'Value' : [] };
    
    # Case 1: str val
    if search_type is str:
        for search_val_i in search_targets:
            for row_ii in sheet.iter_rows(values_only=False):
                for cell_iii in row_ii:
                    if (cell_iii.value is None or isinstance(cell_iii.value, numbers.Number) or isinstance(cell_iii.value, datetime.time) or isinstance(cell_iii.value, datetime.datetime)):
                        continue;
                    if (cell_iii.value == search_val_i):
                        if return_first_hit:
                            return cell_iii.coordinate, cell_iii.value;
                        equal_targets['Coordinate'].append(cell_iii.coordinate);
                        equal_targets['Value'].append(cell_iii.value);
                    try:
                        if is_similar(search_val_i, str(cell_iii.value), threshold=threshold):
                            similar_target_hits['Coordinate'].append(cell_iii.coordinate);
                            similar_target_hits['Value'].append(cell_iii.value);
                    except SyntaxError:
                        continue;    
        if return_all_vals:
            return {'Coordinate' : equal_targets['Coordinate'] + similar_target_hits['Coordinate'], 
                   'Value' : equal_targets['Value'] + similar_target_hits['Value']};
        if equal_targets['Coordinate']:  # check if equal hits are on;
            return equal_targets;    
        return (None, None); # return nothing if there are no results 


    
    # Case 2: number
    elif isinstance(search_targets[0], numbers.Number):    
        for search_val_i in search_targets:    
            round_of_number = count_significant_digits(str(search_val_i));       # calculate signifigant figures, value needs to be string
            for row_ii in sheet.iter_rows(values_only=False):
                for cell_iii in row_ii:
                    if not isinstance(cell_iii.value, numbers.Number):
                        continue; # skip redundant values
                    if (round(cell_iii.value, round_of_number) == round(search_val_i, round_of_number)):
                        if return_first_hit:
                            return cell_iii.coordinate, cell_iii.value
                        equal_targets['Coordinate'].append(cell_iii.coordinate);
                        equal_targets['Value'].append(cell_iii.value); 
                    else:
                        similar_target_hits['Coordinate'].append(cell_iii.coordinate);
                        similar_target_hits['Value'].append(cell_iii.value);

        if return_all_vals:
            return {'Coordinate' : equal_targets['Coordinate'] + similar_target_hits['Coordinate'], 
                   'Value' : equal_targets['Value'] + similar_target_hits['Value']};
        elif equal_targets['Value']:
            return equal_targets['Coordinate'];
        else:
            return (None, None);

    
    # Case 3: date or datetime
    elif (search_type is datetime.date) or (search_type is datetime.datetime):
        for search_val_i in search_targets:
            for row_ii in sheet.iter_rows(values_only=False):
                for cell_iii in row_ii:
                    if not isinstance(cell_iii.value, datetime.datetime) and not isinstance(cell_iii.value, datetime.date):
                        continue;
                    elif (cell_iii.value == search_val_i):  # check wether we find a direct hit;
                        if return_first_hit:
                            return cell_iii.coordinate, cell_iii.value
                        equal_targets['Coordinate'].append(cell_iii.coordinate);
                        equal_targets['Value'].append(cell_iii.value);  
                    else:
                        similar_target_hits['Value'].append(cell_iii.value);
                        similar_target_hits['Coordinate'].append(cell_iii.coordinate);

        if return_all_vals:
            return {'Coordinate' : equal_targets['Coordinate'] + similar_target_hits['Coordinate'], 
                   'Value' : equal_targets['Value'] + similar_target_hits['Value']};
        elif equal_targets['Value']:
            return equal_targets;
        else:
            return (None, None);
                
    elif (search_type is datetime.time):
        for search_val_i in search_targets:
            for row_ii in sheet.iter_rows(values_only=False):
                for cell_iii in row_ii:
                    if not isinstance(cell_iii.value, datetime.time):
                        continue;
                    if (cell_iii.value == search_val_i):
                        if return_first_hit:
                            return cell_iii.coordinate, cell_iii.value;
                        equal_targets['Coordinate'].append(cell_iii.coordinate);
                        equal_targets['Value'].append(cell_iii.value);  
                    else:
                        similar_target_hits['Value'].append(cell_iii.value);
                        similar_target_hits['Coordinate'].append(cell_iii.coordinate);
        if equal_targets['Value'] and not return_all_vals:
            return equal_targets;     
        elif return_all_vals:
            return {'Coordinate' : equal_targets['Coordinate'] + similar_target_hits['Coordinate'], 
                   'Value' : equal_targets['Value'] + similar_target_hits['Value']};
                    
    else:
        raise ValueError(f"Variable search_targets data type unknown : {search_type}, expected str, int, float, datetime.date, datetime.datetime")








def get_excl_sheet_vals(sheet : Worksheet, 
                       start_Cord : tuple[int,int] | str, 
                       end_Cord :   tuple[int,int] | str):
    """
    Takes a given excel sheet and extracts, and returns the values from the start and end coordinates in a 2D array.
    """
    from numpy import array
    # Test data type for start values, ensure the data type is either str or tuple[in,int]
    if isinstance(start_Cord, str):
        row_start, col_start = exc_coord_to_rc(start_Cord);
    elif isinstance(start_Cord, tuple):
        if isinstance(start_Cord[0], int) and isinstance(start_Cord[1], int):
            row_start, col_start = start_Cord;
    else:
        raise TypeError(f"Expected type 'tuple[int, int]' or 'str' for 'start_Cord' - recieved '{start_Cord}' : {type(start_Cord)}");

    # Test data type for end values, ensure the data type is either str or tuple[in,int]
    if isinstance(end_Cord, str):
        row_end, col_end = exc_coord_to_rc(end_Cord);
    elif isinstance(end_Cord, tuple):
        if isinstance(end_Cord[0], int) and isinstance(end_Cord[1], int):
            row_end, col_end = end_Cord;
    else:
        raise TypeError(f"Expected type 'tuple[int, int]' or 'str' for 'end_Cord' - recieved '{end_Cord}' : {type(end_Cord)}");

    # Ensure starting values are lower than ending values for rows and cols
    if row_start > row_end:
        buffer_val = row_end;   # place holder
        row_end = row_start;    # switch values
        row_start = buffer_val; # switch col 
        
    if col_start > col_end:
        buffer_val =  col_end;
        col_end = col_start;
        col_start = buffer_val;
    
    values_extracted = [];
    
    for r_i in range(row_start, row_end):
        values_extracted.append(array([sheet.cell(row=r_i, column = c_ii).value for c_ii in range(col_start, col_end+1)]));

    if (row_start == row_end) or (col_start == col_end):
        return array(values_extracted).ravel()
    return array(values_extracted);       